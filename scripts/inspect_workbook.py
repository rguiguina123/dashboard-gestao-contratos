from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def score_header(values: list[object]) -> int:
    keywords = {
        "nome", "cpf", "função", "funcao", "posto", "sec", "secretaria",
        "contrato", "vigência", "vigencia", "valor", "estado", "uf", "custo",
        "situação", "situacao", "fornecedor", "objeto", "vencimento"
    }
    strings = [str(value).strip().lower() for value in values if isinstance(value, str) and value.strip()]
    return sum(1 for value in strings if any(keyword in value for keyword in keywords)) * 10 + len(strings)


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Uso: inspect_workbook.py <arquivo.xlsx>")

    source = Path(sys.argv[1])
    workbook = load_workbook(source, read_only=True, data_only=True)
    report: list[dict[str, object]] = []

    for sheet in workbook.worksheets:
        candidates: list[tuple[int, int, list[object]]] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True), start=1):
            values = [value for value in row if value is not None and str(value).strip()]
            if values:
                candidates.append((score_header(values), row_number, values))

        candidates.sort(reverse=True, key=lambda item: item[0])
        best = candidates[0] if candidates else (0, None, [])
        report.append({
            "sheet": sheet.title,
            "rows": sheet.max_row,
            "columns": sheet.max_column,
            "detected_header_row": best[1],
            "detected_headers": best[2],
        })

    print(json.dumps({"file": source.name, "sheets": report}, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
